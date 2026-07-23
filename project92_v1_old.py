import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

def process_files(file_paths, tolerance):
    dfs = []
    for i, path in enumerate(file_paths, 1):
        df = pd.read_csv(path, header=None)
        if list(df.iloc[0]).count('pin') or list(df.iloc[0]).count('x'):
            df.columns = df.iloc[0]
            df = df[1:].reset_index(drop=True)
        else:
            df.columns = ['pin', 'condition', 'x', 'y', 'c', 'd']
        df = df[['pin', 'x', 'y', 'c', 'd']].copy()
        df['pin'] = df['pin'].astype(str)
        df.columns = ['pin', f'x{i}', f'y{i}', f'c{i}', f'd{i}']
        dfs.append(df)

    merged_df = dfs[0]
    for df in dfs[1:]:
        merged_df = merged_df.merge(df, on='pin', how='inner')

    for i in range(1, len(dfs) + 1):
        merged_df[f'x{i}_microns'] = pd.to_numeric(merged_df[f'x{i}']) * 1000
        merged_df[f'y{i}_microns'] = pd.to_numeric(merged_df[f'y{i}']) * 1000
        merged_df[f'c{i}_microns'] = pd.to_numeric(merged_df[f'c{i}']) * 1000
        merged_df[f'd{i}_microns'] = pd.to_numeric(merged_df[f'd{i}']) * 1000

    x_cols = [f'x{i}_microns' for i in range(1, len(dfs) + 1)]
    y_cols = [f'y{i}_microns' for i in range(1, len(dfs) + 1)]

    merged_df['range_x'] = merged_df[x_cols].max(axis=1) - merged_df[x_cols].min(axis=1)
    merged_df['range_y'] = merged_df[y_cols].max(axis=1) - merged_df[y_cols].min(axis=1)
    merged_df['fail_x'] = merged_df['range_x'] > tolerance
    merged_df['fail_y'] = merged_df['range_y'] > tolerance
    merged_df['fail_any'] = merged_df['fail_x'] | merged_df['fail_y']

    # Save Excel
    today = datetime.today().strftime('%Y-%m-%d')
    output_file = f'repeatability_results_{today}.xlsx'

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        merged_df[['pin', 'range_x', 'range_y', 'fail_x', 'fail_y']].to_excel(writer, index=False, sheet_name='Results')

        summary = pd.DataFrame({
            'Metric': [
                'Number of tests',
                'Total pins',
                f'Pins failing X (>{tolerance}μm)',
                f'Pins failing Y (>{tolerance}μm)',
                f'Number of pins failing (>{tolerance}μm)',
                'Failing pins number'
            ],
            'Value': [
                len(dfs),
                len(merged_df),
                merged_df['fail_x'].sum(),
                merged_df['fail_y'].sum(),
                merged_df['fail_any'].sum(),
                ', '.join(merged_df[merged_df['fail_any']]['pin'].astype(str).tolist())
            ]
        })
        summary.to_excel(writer, index=False, sheet_name='Summary')

    # Conditional formatting
    wb = openpyxl.load_workbook(output_file)
    ws = wb['Results']
    green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    for col in ['D', 'E']:
        ws.conditional_formatting.add(
            f'{col}2:{col}{ws.max_row}',
            FormulaRule(formula=[f'{col}2=TRUE'], stopIfTrue=True, fill=green)
        )
        ws.conditional_formatting.add(
            f'{col}2:{col}{ws.max_row}',
            FormulaRule(formula=[f'{col}2=FALSE'], stopIfTrue=True, fill=red)
        )

    wb.save(output_file)
    wb.close()

    messagebox.showinfo("Done", f"Report saved as {output_file}")

def start_gui():
    root = tk.Tk()
    root.title("Project 92 Summary Generator")

    tk.Label(root, text="Tolerance (μm):").pack(pady=5)
    tolerance_entry = tk.Entry(root)
    tolerance_entry.insert(0, "2")
    tolerance_entry.pack(pady=5)

    def on_select_files():
        file_paths = filedialog.askopenfilenames(filetypes=[("CSV files", "*.csv")])
        if file_paths:
            try:
                tol = float(tolerance_entry.get())
                process_files(file_paths, tol)
            except ValueError:
                messagebox.showerror("Invalid input", "Please enter a valid number for tolerance.")

    tk.Button(root, text="Upload CSVs and Generate Report", command=on_select_files).pack(pady=10)
    root.mainloop()

if __name__ == '__main__':
    start_gui()
