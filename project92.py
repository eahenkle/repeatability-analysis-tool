import pandas as pd
import numpy as np
import tkinter as tk
from tkinter import filedialog, messagebox
import os
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import subprocess
import platform

# GUI SETUP
root = tk.Tk()
root.title("Project 92 - Repeatability Checker")
root.geometry("500x250")

def upload_files():
    file_paths = filedialog.askopenfilenames(filetypes=[("CSV files", "*.csv")])
    if not file_paths:
        return
    file_listbox.delete(0, tk.END)
    for path in file_paths:
        file_listbox.insert(tk.END, path)
    status_label.config(text="✔️ 檔案已選擇")

def generate_report():
    try:
        files = file_listbox.get(0, tk.END)
        if len(files) < 2:
            messagebox.showwarning("警告", "請至少選擇兩個檔案")
            return

        tolerance = float(tolerance_entry.get())

        dfs = []
        for idx, path in enumerate(files, 1):
            df = pd.read_csv(path, header=None)
            # Handle CSV without headers
            df.columns = ['pin', 'condition', 'x', 'y', 'c', 'd']
            df = df[['pin', 'x', 'y']].copy()
            df['pin'] = df['pin'].astype(str)
            df.columns = ['pin', f'x{idx}', f'y{idx}']
            dfs.append(df)

        merged_df = dfs[0]
        for df in dfs[1:]:
            merged_df = merged_df.merge(df, on='pin', how='inner')

        for i in range(1, len(dfs)+1):
            merged_df[f'x{i}'] = pd.to_numeric(merged_df[f'x{i}']) * 1000
            merged_df[f'y{i}'] = pd.to_numeric(merged_df[f'y{i}']) * 1000

        x_cols = [f'x{i}' for i in range(1, len(dfs)+1)]
        y_cols = [f'y{i}' for i in range(1, len(dfs)+1)]
        merged_df['range_x'] = merged_df[x_cols].max(axis=1) - merged_df[x_cols].min(axis=1)
        merged_df['range_y'] = merged_df[y_cols].max(axis=1) - merged_df[y_cols].min(axis=1)

        merged_df['result_x'] = np.where(merged_df['range_x'] > tolerance, 'NG', 'OK')
        merged_df['result_y'] = np.where(merged_df['range_y'] > tolerance, 'NG', 'OK')
        merged_df['fail_any'] = (merged_df['result_x'] == 'NG') | (merged_df['result_y'] == 'NG')

        # Prepare summary table
        summary = pd.DataFrame({
            'Metric': [
                'Number of tests',
                'Total pins',
                f'Pins failing X (>{tolerance}μm)',
                f'Pins failing Y (>{tolerance}μm)',
                f'Number of pins failing (>{tolerance}μm)',
                'Failing pins number'
            ],
            'Value': [
                len(dfs),
                len(merged_df),
                (merged_df['result_x'] == 'NG').sum(),
                (merged_df['result_y'] == 'NG').sum(),
                merged_df['fail_any'].sum(),
                ', '.join(merged_df[merged_df['fail_any']]['pin'].astype(str).tolist())
            ]
        })

        folder_path = filedialog.askdirectory(title="選擇儲存位置")
        if not folder_path:
            return
        now = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(folder_path, f"repeatability_results_{now}.xlsx")

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            merged_df[['pin', 'range_x', 'range_y', 'result_x', 'result_y']].to_excel(writer, index=False, sheet_name='Results')
            summary.to_excel(writer, index=False, sheet_name='Summary')

        # Conditional formatting
        wb = openpyxl.load_workbook(output_path)
        ws = wb['Results']
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=5):
            for cell in row:
                if cell.value == "OK":
                    cell.fill = green
                elif cell.value == "NG":
                    cell.fill = red

        wb.save(output_path)
        wb.close()
        messagebox.showinfo("成功儲存", f"已將結果儲存至：\n{output_path}")

        # Ask to open the report
        if messagebox.askyesno("開啟報告", "是否要開啟生成的報告？"):
            if platform.system() == "Windows":
                os.startfile(output_path)
            else:
                subprocess.run(["open", output_path] if platform.system() == "Darwin" else ["xdg-open", output_path])

    except Exception as e:
        messagebox.showerror("錯誤", f"處理時發生錯誤：\n{str(e)}")

frame = tk.Frame(root)
frame.pack(pady=10)
upload_button = tk.Button(frame, text="📂 上傳 CSV", command=upload_files)
upload_button.grid(row=0, column=0, padx=10)
tk.Label(frame, text="容許範圍 (μm):").grid(row=0, column=1)
tolerance_entry = tk.Entry(frame, width=10)
tolerance_entry.insert(0, "2")
tolerance_entry.grid(row=0, column=2)
generate_button = tk.Button(frame, text="⚙️ 產生報告", command=generate_report)
generate_button.grid(row=0, column=3, padx=10)

file_listbox = tk.Listbox(root, height=5, width=70)
file_listbox.pack(pady=10)
status_label = tk.Label(root, text="等待上傳...", fg="gray")
status_label.pack()

root.mainloop()